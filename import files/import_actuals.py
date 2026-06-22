"""
Import actuals from 'data files (1).xlsx' into sc_kpi_history table.
Since sc_kpis.code may be empty (old import), we match by POSITION within each scorecard.
Both files have 292 KPIs in the same order.
"""

import openpyxl
import psycopg2
from datetime import datetime
import sys

# ===================== CONFIG =====================
DB_HOST = "localhost"
DB_PORT = 5432
DB_NAME = "stratroom"
DB_USER = "postgres"
DB_PASS = "1234"
DB_SCHEMA = "orgstructure"

SCORECARD_FILE = r"C:\Users\sibi\Desktop\poststratroom\Stratroom-Source\Stratroom-Source\import files\StratRoom_Scorecards_Updated.xlsx"
DATA_FILE = r"C:\Users\sibi\Desktop\poststratroom\Stratroom-Source\Stratroom-Source\import files\data files (1).xlsx"
# ==================================================

def parse_date(s):
    if not s:
        return None
    for fmt in ["%b %d, %Y", "%Y-%m-%d", "%m/%d/%Y"]:
        try:
            return datetime.strptime(str(s).strip(), fmt).date()
        except:
            pass
    return None

def parse_num(v):
    if v is None:
        return None
    try:
        return float(str(v).replace(",", "").replace("%", "").strip())
    except:
        return None

# Step 1: Build ordered KPI lists from both files
print("Building KPI mapping...")
wb_sc = openpyxl.load_workbook(SCORECARD_FILE, read_only=True)
ws_sc = wb_sc[wb_sc.sheetnames[0]]
sc_kpi_codes = []  # ordered: [kpi_code_L0_Board-P1.O1.K1, ...]
seen = set()
for i, row in enumerate(ws_sc.iter_rows(values_only=True)):
    if i == 0:
        continue
    if row[12] and str(row[12]).strip() not in seen:
        seen.add(str(row[12]).strip())
        sc_kpi_codes.append(str(row[12]).strip())

wb_d = openpyxl.load_workbook(DATA_FILE, read_only=True)
ws_d = wb_d['SMVAF_CSE_KRI_ETL_2026']
d_node_keys = []  # ordered: [STRATROOM-BOARD-K01, ...]
seen_d = set()
for i, row in enumerate(ws_d.iter_rows(values_only=True)):
    if i == 0:
        continue
    if row[0] and str(row[0]).strip() not in seen_d:
        seen_d.add(str(row[0]).strip())
        d_node_keys.append(str(row[0]).strip())

node_to_code = {}
for i, nk in enumerate(d_node_keys):
    if i < len(sc_kpi_codes):
        node_to_code[nk] = sc_kpi_codes[i]

print(f"  Mapped {len(node_to_code)} node keys to KPI codes")

# Step 2: Connect to DB
print("Connecting to database...")
conn = psycopg2.connect(
    host=DB_HOST, port=DB_PORT, dbname=DB_NAME,
    user=DB_USER, password=DB_PASS,
    options=f"-c search_path={DB_SCHEMA}"
)
cur = conn.cursor()

# Try matching by code first
cur.execute("SELECT id, code FROM sc_kpis WHERE is_deleted = false AND code IS NOT NULL AND code != ''")
code_to_id = {row[1].strip(): row[0] for row in cur.fetchall()}
print(f"  Loaded {len(code_to_id)} KPI codes from DB")

# If no codes in DB, match by POSITION (order within sc_kpis table)
if len(code_to_id) == 0:
    print("  No codes in DB - using position-based matching...")
    cur.execute("""
        SELECT k.id FROM sc_kpis k
        JOIN sc_objectives o ON k.objective_id = o.id
        JOIN sc_perspectives p ON o.perspective_id = p.id
        JOIN sc_scorecards s ON p.scorecard_id = s.id
        WHERE k.is_deleted = false
        ORDER BY s.id, p.display_order, o.display_order, k.display_order, k.id
    """)
    kpi_ids_ordered = [row[0] for row in cur.fetchall()]
    print(f"  Loaded {len(kpi_ids_ordered)} KPI IDs from DB (position-based)")
    
    # Map node_key -> kpi_id directly by position
    node_to_db_id = {}
    for i, nk in enumerate(d_node_keys):
        if i < len(kpi_ids_ordered):
            node_to_db_id[nk] = kpi_ids_ordered[i]
    
    # Also backfill codes in DB
    print("  Backfilling KPI codes in DB...")
    backfilled = 0
    for i, nk in enumerate(d_node_keys):
        if i < len(sc_kpi_codes) and i < len(kpi_ids_ordered):
            cur.execute("UPDATE sc_kpis SET code = %s WHERE id = %s AND (code IS NULL OR code = '')",
                       (sc_kpi_codes[i], kpi_ids_ordered[i]))
            backfilled += cur.rowcount
    conn.commit()
    print(f"  Backfilled {backfilled} KPI codes")
else:
    # Build node -> db_id via code mapping
    node_to_db_id = {}
    for nk, code in node_to_code.items():
        db_id = code_to_id.get(code)
        if db_id:
            node_to_db_id[nk] = db_id

print(f"  Ready to match {len(node_to_db_id)} node keys to DB KPI IDs")

# Step 3: Read data file and insert actuals
print("Importing actuals into sc_kpi_history...")
wb_d2 = openpyxl.load_workbook(DATA_FILE, read_only=True)
ws_d2 = wb_d2['SMVAF_CSE_KRI_ETL_2026']

inserted = 0
skipped = 0
errors = []

for i, row in enumerate(ws_d2.iter_rows(values_only=True)):
    if i == 0:
        continue

    node_key = str(row[0]).strip() if row[0] else None
    period_start = parse_date(row[2])
    period_end = parse_date(row[3])
    target = parse_num(row[4])
    actual = parse_num(row[5])

    if not node_key or not period_start or not period_end:
        skipped += 1
        continue

    kpi_id = node_to_db_id.get(node_key)
    if not kpi_id:
        errors.append(f"No DB match for node_key: {node_key}")
        skipped += 1
        continue

    if actual is None and target is None:
        skipped += 1
        continue

    try:
        cur.execute("""
            INSERT INTO sc_kpi_history (kpi_id, period_start, period_end, actual_value)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (kpi_id, period_start, period_end)
            DO UPDATE SET
                actual_value = COALESCE(EXCLUDED.actual_value, sc_kpi_history.actual_value),
                calculated_at = NOW()
        """, (kpi_id, period_start, period_end, actual))
        inserted += 1
    except Exception as e:
        errors.append(f"Row {i}: {e}")
        conn.rollback()
        skipped += 1
        continue

conn.commit()
cur.close()
conn.close()

print(f"\nImport complete!")
print(f"  Inserted/Updated: {inserted}")
print(f"  Skipped: {skipped}")
if errors:
    print(f"\nErrors ({len(errors)}):")
    for e in errors[:20]:
        print(f"  - {e}")
else:
    print("  No errors!")
