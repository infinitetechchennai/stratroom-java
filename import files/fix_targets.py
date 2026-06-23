"""
Backfill sc_kpis.target_value from the data file's Target column.
Also re-imports actuals if needed.
"""

import openpyxl
import psycopg2
from datetime import datetime

# ===================== CONFIG =====================
DB_HOST = "localhost"
DB_PORT = 5432
DB_NAME = "stratroom"
DB_USER = "postgres"
DB_PASS = "1234"
DB_SCHEMA = "orgstructure"

SCORECARD_FILE = r"C:\Users\sibi\Desktop\poststratroom\Stratroom-Source\Stratroom-Source\import files\StratRoom_Scorecards_Updated.xlsx"
DATA_FILE = r"C:\Users\sibi\Desktop\poststratroom\Stratroom-Source\Stratroom-Source\import files\data files (1).xlsx"
# ==================================================

def parse_num(v):
    if v is None:
        return None
    try:
        return float(str(v).replace(",", "").replace("%", "").strip())
    except:
        return None

print("Connecting to database...")
conn = psycopg2.connect(
    host=DB_HOST, port=DB_PORT, dbname=DB_NAME,
    user=DB_USER, password=DB_PASS,
    options=f"-c search_path={DB_SCHEMA}"
)
cur = conn.cursor()

# Load code -> kpi_id from DB
cur.execute("SELECT id, code FROM sc_kpis WHERE is_deleted = false AND code IS NOT NULL AND code != ''")
code_to_id = {row[1].strip(): row[0] for row in cur.fetchall()}
print(f"  Loaded {len(code_to_id)} KPI codes from DB")

# Load ordered KPI codes from scorecard file (position mapping)
wb_sc = openpyxl.load_workbook(SCORECARD_FILE, read_only=True)
ws_sc = wb_sc[wb_sc.sheetnames[0]]
sc_kpi_codes = []
seen = set()
for i, row in enumerate(ws_sc.iter_rows(values_only=True)):
    if i == 0:
        continue
    if row[12] and str(row[12]).strip() not in seen:
        seen.add(str(row[12]).strip())
        sc_kpi_codes.append(str(row[12]).strip())

# Load ordered node keys from data file
wb_d = openpyxl.load_workbook(DATA_FILE, read_only=True)
ws_d = wb_d['SMVAF_CSE_KRI_ETL_2026']
d_node_keys = []
seen_d = set()
for i, row in enumerate(ws_d.iter_rows(values_only=True)):
    if i == 0:
        continue
    if row[0] and str(row[0]).strip() not in seen_d:
        seen_d.add(str(row[0]).strip())
        d_node_keys.append(str(row[0]).strip())

# Build node_key -> kpi_code mapping
node_to_code = {}
for i, nk in enumerate(d_node_keys):
    if i < len(sc_kpi_codes):
        node_to_code[nk] = sc_kpi_codes[i]

# Step 1: Backfill target_value from data file into sc_kpis
# Use the first occurrence of each node_key to get its target
print("Backfilling target values from data file into sc_kpis...")
node_to_target = {}
wb_d2 = openpyxl.load_workbook(DATA_FILE, read_only=True)
ws_d2 = wb_d2['SMVAF_CSE_KRI_ETL_2026']
for i, row in enumerate(ws_d2.iter_rows(values_only=True)):
    if i == 0:
        continue
    node_key = str(row[0]).strip() if row[0] else None
    if node_key and node_key not in node_to_target:
        t = parse_num(row[4])  # Target column
        if t is not None:
            node_to_target[node_key] = t

updated_targets = 0
for node_key, target in node_to_target.items():
    kpi_code = node_to_code.get(node_key)
    if not kpi_code:
        continue
    kpi_id = code_to_id.get(kpi_code)
    if not kpi_id:
        continue
    cur.execute(
        "UPDATE sc_kpis SET target_value = %s WHERE id = %s",
        (target, kpi_id)
    )
    updated_targets += cur.rowcount

conn.commit()
print(f"  Updated target_value for {updated_targets} KPIs")

cur.close()
conn.close()
print("\nDone! Refresh your scorecard page - targets should now show correctly.")
